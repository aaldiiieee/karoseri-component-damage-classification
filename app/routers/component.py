import logging
from typing import Optional
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from ..configs.security import get_current_user
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.responses import StreamingResponse
from app.configs.db import get_db
from ..schemas import (
    ComponentCreate,
    ComponentUpdate,
    ComponentResponse,
    ComponentList
)
from ..schemas.damage_record import BulkImportResult
from app.services import component_service

logger = logging.getLogger("app")

router = APIRouter(prefix="/components", tags=["Components"])

@router.post("/", response_model=ComponentResponse, status_code=201)
async def create_component(
    data: ComponentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: AsyncSession = Depends(get_current_user)
):
    """Create a new component."""
    if current_user.role not in ["superadmin", "admin"]:
        raise HTTPException(
            status_code=403,
            detail="Anda tidak memiliki hak akses untuk mengakses resource ini"
        )

    # Check if component code already exists
    existing = await component_service.get_by_code(db, data.code)
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Component with code '{data.code}' already exists"
        )
    
    component = await component_service.create(db, data)
    return component


@router.get("/", response_model=ComponentList)
async def get_components(
    page: int = Query(1, ge=1, description="Page number"),
    size: int = Query(10, ge=1, le=100, description="Items per page"),
    category: Optional[str] = Query(None, description="Filter by category"),
    is_active: Optional[bool] = Query(None, description="Filter by active status"),
    search: Optional[str] = Query(None, description="Search by code or name"),
    db: AsyncSession = Depends(get_db),
    current_user: AsyncSession = Depends(get_current_user)
):
    """Get all components with pagination and filters."""
    if current_user.role not in ["superadmin", "admin"]:
        raise HTTPException(
            status_code=403,
            detail="Anda tidak memiliki hak akses untuk mengakses resource ini"
        )

    items, total = await component_service.get_all(
        db, page=page, size=size, category=category, is_active=is_active, search=search
    )
    
    pages = (total + size - 1) // size  # Ceiling division
    
    return ComponentList(
        items=items,
        total=total,
        page=page,
        size=size,
        pages=pages
    )


@router.get("/categories", response_model=list[str])
async def get_categories(db: AsyncSession = Depends(get_db)):
    """Get all unique component categories."""
    return await component_service.get_categories(db)


@router.get("/import-template")
async def download_component_import_template(current_user: AsyncSession = Depends(get_current_user)):
    """Download an Excel template for bulk importing components."""
    if current_user.role not in ["superadmin", "admin"]:
        raise HTTPException(
            status_code=403,
            detail="Anda tidak memiliki hak akses untuk mengakses resource ini"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Template Komponen"
    
    headers = ["code", "name", "category", "description"]
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    examples = [
        ["KRS-001", "Roof Panel", "Body Panel", "Panel atap kendaraan"],
        ["KRS-002", "Side Panel", "Body Panel", "Panel samping kendaraan"],
        ["KRS-003", "Floor Panel", "Chassis", "Panel lantai kendaraan"],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    for col in ws.columns:
        max_length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_komponen.xlsx"}
    )


@router.post("/bulk-import", response_model=BulkImportResult)
async def bulk_import_components(
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV file with components"),
    db: AsyncSession = Depends(get_db),
    current_user: AsyncSession = Depends(get_current_user)
):
    if current_user.role not in ["superadmin", "admin"]:
        raise HTTPException(
            status_code=403,
            detail="Anda tidak memiliki hak akses untuk mengakses resource ini"
        )

    """
    Bulk import components from Excel (.xlsx) or CSV file.
    
    Expected columns: code, name, category, description
    """
    filename = file.filename.lower()
    
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(
            status_code=400,
            detail="Format file tidak didukung. Gunakan file .xlsx atau .csv"
        )
    
    content = await file.read()
    rows = []
    
    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                headers.append(str(cell.value).strip().lower() if cell.value else "")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                row_dict = {}
                for idx, header in enumerate(headers):
                    row_dict[header] = row[idx] if idx < len(row) else None
                rows.append(row_dict)
            
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Gagal membaca file Excel: {str(e)}")
    else:
        decoded = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = list(reader)
    
    records = []
    errors = []
    
    for i, row in enumerate(rows, start=2):
        try:
            record_data = ComponentCreate(
                code=str(row.get("code", "")).strip(),
                name=str(row.get("name", "")).strip(),
                category=str(row.get("category", "")).strip(),
                description=str(row.get("description", "")).strip() or None,
            )
            records.append(record_data)
        except Exception as e:
            errors.append(f"Baris {i}: {str(e)}")
    
    if records:
        success, error_count, bulk_errors = await component_service.bulk_create(db, records)
        errors.extend(bulk_errors)
    else:
        success = 0
    
    return BulkImportResult(
        success_count=success,
        error_count=len(errors),
        errors=errors[:20]
    )


@router.get("/{component_id}", response_model=ComponentResponse)
async def get_component(
    component_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: AsyncSession = Depends(get_current_user)
):
    """Get a component by ID."""
    if current_user.role not in ["superadmin", "admin"]:
        raise HTTPException(
            status_code=403,
            detail="Anda tidak memiliki hak akses untuk mengakses resource ini"
        )
        
    component = await component_service.get_by_id(db, component_id)
    if not component:
        raise HTTPException(status_code=404, detail="Component not found")
    return component


@router.put("/{component_id}", response_model=ComponentResponse)
async def update_component(
    component_id: UUID,
    data: ComponentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: AsyncSession = Depends(get_current_user)
):
    """Update a component."""
    if current_user.role not in ["superadmin", "admin"]:
        raise HTTPException(
            status_code=403,
            detail="Anda tidak memiliki hak akses untuk mengakses resource ini"
        )

    # Check if new code already exists (if updating code)
    if data.code:
        existing = await component_service.get_by_code(db, data.code)
        if existing and existing.id != component_id:
            raise HTTPException(
                status_code=400,
                detail=f"Component with code '{data.code}' already exists"
            )
    
    component = await component_service.update(db, component_id, data)
    if not component:
        raise HTTPException(status_code=404, detail="Component not found")
    return component


@router.delete("/{component_id}", status_code=204)
async def delete_component(
    component_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: AsyncSession = Depends(get_current_user)
):
    """Delete a component."""
    if current_user.role not in ["superadmin", "admin"]:
        raise HTTPException(
            status_code=403,
            detail="Anda tidak memiliki hak akses untuk mengakses resource ini"
        )
        
    deleted = await component_service.delete(db, component_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Component not found")
    return None