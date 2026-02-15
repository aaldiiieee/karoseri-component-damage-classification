import logging
from typing import Optional
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
import csv
import io

from ..configs.db import get_db
from ..schemas import (
    DamageRecordCreate,
    DamageRecordUpdate,
    DamageRecordResponse,
    DamageRecordList,
    DamageDistribution,
    BulkImportResult
)
from ..services import damage_record_service, component_service

logger = logging.getLogger("app")

router = APIRouter(prefix="/damage-records", tags=["Damage Records"])


@router.post("/", response_model=DamageRecordResponse, status_code=201)
async def create_damage_record(
    data: DamageRecordCreate,
    db: AsyncSession = Depends(get_db)
):
    """Create a new damage record (training data)."""
    # Verify component exists
    component = await component_service.get_by_id(db, data.component_id)
    if not component:
        raise HTTPException(status_code=404, detail="Component not found")
    
    record = await damage_record_service.create(db, data)
    return record


@router.get("/", response_model=DamageRecordList)
async def get_damage_records(
    page: int = Query(1, ge=1, description="Page number"),
    size: int = Query(10, ge=1, le=100, description="Items per page"),
    component_id: Optional[UUID] = Query(None, description="Filter by component"),
    damage_level: Optional[str] = Query(None, description="Filter by damage level (ringan/sedang/berat)"),
    db: AsyncSession = Depends(get_db)
):
    """Get all damage records with pagination and filters."""
    # Validate damage_level
    if damage_level:
        damage_level = damage_level.lower()
        if damage_level not in ["ringan", "sedang", "berat"]:
            raise HTTPException(
                status_code=400,
                detail="damage_level must be one of: ringan, sedang, berat"
            )
    
    items, total = await damage_record_service.get_all(
        db, page=page, size=size, component_id=component_id, damage_level=damage_level
    )
    
    pages = (total + size - 1) // size
    
    return DamageRecordList(
        items=items,
        total=total,
        page=page,
        size=size,
        pages=pages
    )


@router.get("/distribution", response_model=DamageDistribution)
async def get_damage_distribution(db: AsyncSession = Depends(get_db)):
    """Get damage level distribution statistics."""
    return await damage_record_service.get_distribution(db)


@router.get("/import-template")
async def download_import_template():
    """Download an Excel template for bulk importing damage records."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Template Data Kerusakan"
    
    headers = [
        "component_code", "damage_area", "damage_depth", "damage_point_count",
        "component_age", "usage_frequency", "corrosion_level", "deformation",
        "damage_level", "notes"
    ]
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"), 
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Example rows
    examples = [
        ["KRS-001", 2.5, 0.5, 2, 6, 3, 1, 0.3, "ringan", "Kerusakan ringan"],
        ["KRS-001", 15.0, 3.0, 5, 24, 7, 3, 2.5, "sedang", "Kerusakan sedang"],
        ["KRS-001", 35.0, 7.0, 10, 48, 9, 5, 6.0, "berat", "Kerusakan berat"],
    ]
    
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 4
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=template_data_kerusakan.xlsx"
        }
    )


@router.get("/{record_id}", response_model=DamageRecordResponse)
async def get_damage_record(
    record_id: UUID,
    db: AsyncSession = Depends(get_db)
):
    """Get a damage record by ID."""
    record = await damage_record_service.get_by_id(db, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Damage record not found")
    return record


@router.put("/{record_id}", response_model=DamageRecordResponse)
async def update_damage_record(
    record_id: UUID,
    data: DamageRecordUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update a damage record."""
    # Verify component exists if updating component_id
    if data.component_id:
        component = await component_service.get_by_id(db, data.component_id)
        if not component:
            raise HTTPException(status_code=404, detail="Component not found")
    
    record = await damage_record_service.update(db, record_id, data)
    if not record:
        raise HTTPException(status_code=404, detail="Damage record not found")
    return record


@router.delete("/{record_id}", status_code=204)
async def delete_damage_record(
    record_id: UUID,
    db: AsyncSession = Depends(get_db)
):
    """Delete a damage record."""
    deleted = await damage_record_service.delete(db, record_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Damage record not found")
    return None


@router.post("/bulk-import", response_model=BulkImportResult)
async def bulk_import_damage_records(
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV file with damage records"),
    db: AsyncSession = Depends(get_db)
):
    """
    Bulk import damage records from Excel (.xlsx) or CSV file.
    
    Expected columns:
    component_code, damage_area, damage_depth, damage_point_count, 
    component_age, usage_frequency, corrosion_level, deformation, damage_level, notes
    """
    filename = file.filename.lower()
    
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(
            status_code=400, 
            detail="Format file tidak didukung. Gunakan file .xlsx atau .csv"
        )
    
    content = await file.read()
    rows = []
    
    if filename.endswith(".xlsx"):
        # Parse Excel file with openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                headers.append(str(cell.value).strip().lower() if cell.value else "")
            
            # Parse data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue  # Skip empty rows
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        row_dict[header] = row[idx]
                    else:
                        row_dict[header] = None
                rows.append(row_dict)
            
            wb.close()
        except Exception as e:
            raise HTTPException(
                status_code=400, 
                detail=f"Gagal membaca file Excel: {str(e)}"
            )
    else:
        # Parse CSV file
        decoded = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = list(reader)
    
    records = []
    errors = []
    
    for i, row in enumerate(rows, start=2):  # Start at 2 (1 for header)
        try:
            # Get component by code
            component_code = str(row.get("component_code", "")).strip()
            component = await component_service.get_by_code(db, component_code)
            
            if not component:
                errors.append(f"Baris {i}: Komponen '{component_code}' tidak ditemukan")
                continue
            
            record_data = DamageRecordCreate(
                component_id=component.id,
                damage_area=float(row["damage_area"]),
                damage_depth=float(row["damage_depth"]),
                damage_point_count=int(row["damage_point_count"]),
                component_age=int(row["component_age"]),
                usage_frequency=int(row["usage_frequency"]),
                corrosion_level=int(row["corrosion_level"]),
                deformation=float(row["deformation"]),
                damage_level=str(row["damage_level"]).strip(),
                notes=str(row.get("notes", "")).strip() or None
            )
            records.append(record_data)
        except Exception as e:
            errors.append(f"Baris {i}: {str(e)}")
    
    # Bulk create valid records
    if records:
        success, error_count, bulk_errors = await damage_record_service.bulk_create(db, records)
        errors.extend(bulk_errors)
    else:
        success = 0
    
    return BulkImportResult(
        success_count=success,
        error_count=len(errors),
        errors=errors[:20]  # Limit to first 20 errors
    )